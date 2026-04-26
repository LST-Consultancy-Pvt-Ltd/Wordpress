"""Generate an Excel file of the app navigation structure.
Changes applied:
- Local SEO group: keep only Programmatic SEO and Keyword Clusters
- Remove entire Off-Page SEO group
- Remove entire Commerce group
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Navigation Structure"

# Styles
header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
group_font = Font(name="Calibri", bold=True, size=11, color="1F2937")
group_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
item_font = Font(name="Calibri", size=10)
thin_border = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

# Headers
headers = ["Group", "Module", "Route Path", "Icon"]
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 32
ws.column_dimensions["D"].width = 20

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# Navigation data (with changes applied)
nav_groups = [
    (None, [
        ("Dashboard", "/", "LayoutDashboard"),
        ("Sites", "/sites", "Globe"),
        ("AI Command", "/ai-command", "Sparkles"),
        ("Autopilot", "/autopilot", "Bot"),
    ]),
    ("Content", [
        ("Pages", "/pages", "FileText"),
        ("Posts", "/posts", "Newspaper"),
        ("Calendar", "/calendar", "CalendarDays"),
        ("Content Refresh", "/content-refresh", "RefreshCw"),
        ("Live Editor", "/live-editor", "PenTool"),
        ("Media Library", "/media-library", "Image"),
        ("Comments", "/comments", "MessageCircle"),
    ]),
    ("WordPress", [
        ("Plugins & Themes", "/plugins-themes", "Puzzle"),
        ("WP Users", "/wp-users", "Users"),
        ("Forms", "/forms", "FileInput"),
        ("Navigation", "/navigation", "Menu"),
        ("Backups", "/backups", "Archive"),
        ("Redirects", "/redirects", "ArrowRightLeft"),
    ]),
    ("SEO & Analytics", [
        ("SEO", "/seo", "Search"),
        ("Search Visibility", "/search-visibility", "Eye"),
        ("Keyword Tracking", "/keyword-tracking", "Target"),
        ("Site Speed", "/site-speed", "Gauge"),
        ("Link Builder", "/link-builder", "Network"),
        ("Broken Links", "/broken-links", "Link2"),
        ("Duplicate Content", "/duplicate-content", "Copy"),
        ("Crawl Report", "/crawl-report", "Bug"),
        ("Local Tracking", "/local-tracking", "MapPin"),
        ("Schema Markup", "/schema-markup", "Code2"),
        ("Sitemap & Robots", "/sitemap-robots", "Map"),
        ("Canonical Manager", "/canonical-manager", "GitMerge"),
        ("Mobile Checker", "/mobile-checker", "Smartphone"),
        ("Reports", "/reports", "BarChart3"),
        ("Report Builder", "/report-builder", "LayoutGrid"),
    ]),
    ("Marketing", [
        ("Social Media", "/social-media", "Share2"),
        ("Newsletter", "/newsletter", "Mail"),
        ("A/B Testing", "/ab-testing", "FlaskConical"),
    ]),
    # Commerce group — REMOVED
    ("Health", [
        ("Site Health", "/site-health", "HeartPulse"),
        ("Activity", "/activity", "Activity"),
    ]),
    # Local SEO — keep ONLY Programmatic SEO and Keyword Clusters
    ("Local SEO", [
        ("Programmatic SEO", "/programmatic-seo", "Layers"),
        ("Keyword Clusters", "/keyword-clusters", "Hash"),
    ]),
    # Off-Page SEO group — REMOVED
    (None, [
        ("Settings", "/settings", "Settings"),
    ]),
]

row = 2
for group_label, items in nav_groups:
    display_group = group_label if group_label else "Core"
    # Group header row
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col)
        cell.fill = group_fill
        cell.font = group_font
        cell.border = thin_border
    ws.cell(row=row, column=1, value=display_group).font = group_font
    ws.cell(row=row, column=1).fill = group_fill
    ws.cell(row=row, column=1).border = thin_border
    row += 1

    for label, path, icon in items:
        ws.cell(row=row, column=1, value="").border = thin_border
        c2 = ws.cell(row=row, column=2, value=label)
        c2.font = item_font
        c2.border = thin_border
        c3 = ws.cell(row=row, column=3, value=path)
        c3.font = item_font
        c3.border = thin_border
        c4 = ws.cell(row=row, column=4, value=icon)
        c4.font = item_font
        c4.border = thin_border
        row += 1

    # Empty separator row
    row += 1

# Freeze top row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:D{row - 1}"

output = r"c:\Users\Sanket.Bute\LST Project\Wordpress\Navigation_Structure.xlsx"
wb.save(output)
print(f"Excel saved to: {output}")
